import os, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = r"D:\AI_Project\Книги"

# ============ 1. COLLECT ALL PDF BOOKS ============
pdf_books = []  # (folder, filename)
for dirpath, dirnames, filenames in os.walk(ROOT):
    dirnames[:] = [d for d in dirnames if not d.startswith('.') and not d.startswith('__')
                   and d not in ('converted', 'converted_new', 'outputs', 'summary', 'temp_docx')]
    folder = os.path.relpath(dirpath, ROOT)
    for f in filenames:
        if f.lower().endswith('.pdf'):
            # Skip analysis and report files
            if f.startswith('анализ_') or f == '_.pdf':
                continue
            pdf_books.append((folder, f))

print(f"Total PDF books: {len(pdf_books)}")

# ============ 2. READING LISTS ============
# Финансы план
finance_plan = [
    "Психологически насыщенная жизнь",
    "Сам себе финансист",
    "Финансы для нефинансистов",
    "Криптвоюматика 2.0",
    "Нескучные финансы",
    "Пора зарабатывать больше!",
    "Пес по имени Мани",
    "Фактор латте",
    "Давай поговорим о твоих доходах и расходах",
    "Девушка с деньгами",
    "Маленькая книга здравого смысла инвестирования",
    "Думай медленно... решай быстро",
    "Деньги без дураков",
    "Психология денег",
    "Случайная прогулка по Уолл-стрит",
    "Разумное распределение активов",
    "Разумный инвестор",
    "Воспоминания биржевого спекулянта",
    "Иррациональный оптимизм",
    "О самом важном",
    "Победа над долгами",
    "Куда уходят деньги",
    "Как сберечь деньги в кризис",
    "Стратегия мышления богатых и бедных",
    "Привычки миллионеров",
    "На пенсию в 35 лет",
    "1000000$ в инвестициях",
    "Секрет обогащения",
    "Психология денег (Чубаров)",
    "Система финансового оздоровления",
    "Магия денег",
    "Освободитесь от плохих долгов",
    "Квадрант денежного потока",
    "Давай поговорим о твоих доходах",
    "Деньги, успех и вы",
    "Путь к финансовой свободе",
    "Пора зарабатывать больше!",
    "Экономим с удовольствием",
    "Ещё одна книга про деньги",
    "Экономика за 30 секунд",
    "Краткая история экономической мысли",
    "Банковский кредит",
    "Телефонных продаж не бывает",
    "Деловое общение",
    "Переговоры без поражения",
    "Анатомия безденежья",
    "Наука завоевывать друзей",
    "Как располагать к себе людей",
    "Дэн Кеннеди",
    "Жесткий тайм-менеджмент",
    "Высокая продуктивность за 30 дней",
    "Минус одна вредная привычка",
    "Мастер Времени",
    "Я умер в понедельник",
    "Топливо на нуле",
    "Биохакинг времени",
    "Не жди понедельника",
    "Агрессия это энергия",
    "Разум чемпиона",
    "Нейроброня",
    "Реальный менеджмент",
    "Как все успевать за 24 часа",
    "Быстрый мозг",
    "Не боюсь работать языком",
    "Цель вашей жизни",
    "Удача, которая всегда с тобой",
    "Быть в Плюсе",
    "Биология добра и зла",
    "Антихрупкость",
    "Атомные привычки",
    "Важные годы",
    "Жить проще, получать больше",
]

# Список нужных книг развитие
dev_plan = [
    "Вся фигня – от мозга?!",
    "зигмунд фрейд введение в психоанали",
    "Сапольски-Роберт.-Психология-стресса",
    "Зоопарк в твоей голове",
    "Джо Диспенза Сила подсознания",
    "Келли Макгонигал Сила воли",
    "Займись ничем: система долгосрочной продуктивности",
    "Обрети утраченную радость",
    "Психология устойчивости",
    "Психологический интеллект",
    "Эмоциональная регуляция",
    "Выбирай себя каждый день",
    "Как сила позитивного мышления сделает вас богатыми",
]

# ============ 3. SEARCH FUNCTION ============
def find_book(books_list, query):
    """Search for a book by keyword match in filenames."""
    q = query.lower().strip()
    # Remove common suffixes
    q = re.sub(r'скачать бесплатно', '', q)
    q = re.sub(r'[\.!\?,:;]', '', q)
    q_words = set(q.split())
    
    results = []
    for folder, fname in books_list:
        fn = fname.lower()
        # Count how many query words match
        matches = sum(1 for w in q_words if len(w) > 3 and w in fn)
        # Check for author name
        author_match = any(a in fn for a in q_words if len(a) > 3)
        if matches >= 2 or author_match:
            # Calculate relevance
            relevance = matches / max(len(q_words), 1)
            results.append((relevance, folder, fname))
    
    results.sort(key=lambda x: -x[0])
    return results

def find_by_exact_name(books_list, query):
    """Try to find book by exact phrase or key terms."""
    q = query.lower().strip()
    # Remove скачать бесплатно
    q = re.sub(r'скачать бесплатно', '', q).strip()
    
    best = None
    best_score = 0
    
    for folder, fname in books_list:
        fn = fname.lower()
        score = 0
        # Check how many significant words match
        q_words = [w for w in re.split(r'[\s\-–—,]+', q) if len(w) > 3]
        for w in q_words:
            if w in fn:
                score += 1
        # Bonus for exact sequence match
        q_short = q[:30]
        if q_short in fn:
            score += 3
            
        if score > best_score:
            best_score = score
            best = (folder, fname)
    
    return best, best_score

# ============ 4. COMPARE LISTS ============
print("\n=== Check Finance Plan ===")
finance_results = []
for book in finance_plan:
    if not book.strip():
        continue
    found, score = find_by_exact_name(pdf_books, book)
    if found and score >= 2:
        folder, fname = found
        finance_results.append((book, "Есть", folder, fname))
        print(f"  ✅ [{folder}] {book[:50]:50s} -> {fname[:40]}")
    else:
        finance_results.append((book, "НЕТ", "", ""))
        print(f"  ❌ {book[:55]}")

print(f"\n=== Check Development Plan ===")
dev_results = []
for book in dev_plan:
    if not book.strip():
        continue
    found, score = find_by_exact_name(pdf_books, book)
    if found and score >= 1:
        folder, fname = found
        dev_results.append((book, "Есть", folder, fname))
        print(f"  ✅ [{folder}] {book[:50]:50s} -> {fname[:40]}")
    else:
        dev_results.append((book, "НЕТ", "", ""))
        print(f"  ❌ {book[:55]}")

# ============ 5. CATEGORIZATION ANALYSIS ============
print("\n=== Folder Analysis ===")
folder_contents = {}
for folder, fname in pdf_books:
    if folder not in folder_contents:
        folder_contents[folder] = []
    folder_contents[folder].append(fname)

# Find misplaced books
misplaced = []
# In Финансы folder - books about sales, mindset
finance_folder = r"Финансы"
for f in folder_contents.get(finance_folder, []):
    fn_lower = f.lower()
    # Check if it belongs elsewhere
    if 'продаж' in fn_lower or 'sales' in fn_lower:
        misplaced.append((finance_folder, f, "Продажи"))
    elif 'мышлен' in fn_lower and 'богат' not in fn_lower and 'денег' not in fn_lower:
        misplaced.append((finance_folder, f, "Мышление, привычки, навыки"))
    elif 'менеджмент' in fn_lower and 'финанс' not in fn_lower:
        misplaced.append((finance_folder, f, "Тайм-менеджмент/Менеджмент"))
    elif 'общени' in fn_lower or 'переговор' in fn_lower or 'оратор' in fn_lower:
        misplaced.append((finance_folder, f, "Общение, красноречие"))

print(f"\nMisplaced books found: {len(misplaced)}")
for folder, fname, target in misplaced:
    print(f"  {folder:15s} -> {target:25s}: {fname[:50]}")

# ============ 6. RECOMMENDED FOLDER STRUCTURE ============
print("\n=== Recommended Folder Structure ===")
categories = {
    "Финансы": "Личные финансы, инвестиции, экономика, богатство",
    "Продажи": "Продажи, переговоры, воронки, клиенты",
    "Психология": "Психология, психика, поведение",
    "Мышление, привычки, навыки": "Мышление, привычки, продуктивность, навыки",
    "Общение, красноречие": "Коммуникация, ораторское искусство, харизма",
    "Тайм-менеджмент": "Тайм-менеджмент, управление временем, продуктивность",
    "Вождение": "Автомобили, ПДД, вождение",
    "Спорт, питание": "Спорт, фитнес, питание, здоровье",
    "свободные_книги": "Классическая экономика (Адам Смит, Кейнс, и др.)",
    "Бизнес, менеджмент": "Управление, лидерство, бизнес-процессы",
    "Эзотерика, саморазвитие": "Эзотерика, чакры, духовные практики, афоризмы",
    "Здоровье, исцеление": "Физическое и ментальное здоровье, исцеление",
    "Философия, психоанализ": "Философия, Фрейд, психоанализ",
}

for cat, desc in categories.items():
    count = 0
    for f, fnames in folder_contents.items():
        if f.endswith(cat) or f == cat:
            count = len(fnames)
            break
    status = "✅" if os.path.isdir(os.path.join(ROOT, cat)) else "➕ Новая"
    print(f"  {status} {cat:30s} ({count:3d} книг) - {desc}")

# ============ 7. CREATE EXCEL ============
print("\n=== Creating Excel ===")
wb = Workbook()

# Sheet 1: Summary
ws1 = wb.active
ws1.title = "Сводка"
ws1.append(["Сводный отчёт по коллекции книг", "", "", ""])
ws1.append(["D:\\AI_Project\\Книги", "", "", ""])
ws1.append([])
ws1.append(["Всего PDF книг:", len(pdf_books), "", ""])
ws1.append(["Папок:", len(folder_contents), "", ""])
ws1.append([])

header_font = Font(bold=True, size=12)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, color="FFFFFF")

# Folders overview
ws1.append(["Структура папок", "Книг", "Рекомендация"])
ws1.cell(row=ws1.max_row, column=1).font = header_font_white
ws1.cell(row=ws1.max_row, column=2).font = header_font_white
ws1.cell(row=ws1.max_row, column=3).font = header_font_white
ws1.cell(row=ws1.max_row, column=1).fill = header_fill
ws1.cell(row=ws1.max_row, column=2).fill = header_fill
ws1.cell(row=ws1.max_row, column=3).fill = header_fill

for folder in sorted(folder_contents.keys()):
    n = len(folder_contents[folder])
    if folder in ('converted', 'converted_new', 'outputs', 'summary', 'temp_docx', '.'):
        continue
    rec = categories.get(folder, "—")
    ws1.append([folder, n, rec])

# Sheet 2: Finance Plan
ws2 = wb.create_sheet("Финансы план")
ws2.append(["Книга из плана", "Статус", "Где найдено", "Файл"])
ws2.cell(row=1, column=1).font = header_font_white
ws2.cell(row=1, column=2).font = header_font_white
ws2.cell(row=1, column=3).font = header_font_white
ws2.cell(row=1, column=4).font = header_font_white
ws2.cell(row=1, column=1).fill = header_fill
ws2.cell(row=1, column=2).fill = header_fill
ws2.cell(row=1, column=3).fill = header_fill
ws2.cell(row=1, column=4).fill = header_fill

red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

for book, status, folder, fname in finance_results:
    row = ws2.max_row + 1
    ws2.append([book, status, folder, fname])
    if status == "НЕТ":
        ws2.cell(row=row, column=2).fill = red_fill
    else:
        ws2.cell(row=row, column=2).fill = green_fill

# Sheet 3: Development Plan
ws3 = wb.create_sheet("Развитие план")
ws3.append(["Книга из плана", "Статус", "Где найдено", "Файл"])
ws3.cell(row=1, column=1).font = header_font_white
ws3.cell(row=1, column=2).font = header_font_white
ws3.cell(row=1, column=3).font = header_font_white
ws3.cell(row=1, column=4).font = header_font_white
ws3.cell(row=1, column=1).fill = header_fill
ws3.cell(row=1, column=2).fill = header_fill
ws3.cell(row=1, column=3).fill = header_fill
ws3.cell(row=1, column=4).fill = header_fill

for book, status, folder, fname in dev_results:
    row = ws3.max_row + 1
    ws3.append([book, status, folder, fname])
    if status == "НЕТ":
        ws3.cell(row=row, column=2).fill = red_fill
    else:
        ws3.cell(row=row, column=2).fill = green_fill

# Sheet 4: Misplaced books
ws4 = wb.create_sheet("Несоответствия")
ws4.append(["Сейчас в папке", "Книга", "Должна быть в"])
ws4.cell(row=1, column=1).font = header_font_white
ws4.cell(row=1, column=2).font = header_font_white
ws4.cell(row=1, column=3).font = header_font_white
ws4.cell(row=1, column=1).fill = header_fill
ws4.cell(row=1, column=2).fill = header_fill
ws4.cell(row=1, column=3).fill = header_fill

for folder, fname, target in misplaced:
    ws4.append([folder, fname[:80], target])

# Sheet 5: All books by category
ws5 = wb.create_sheet("Все книги по папкам")
ws5.append(["Папка", "Файл книги (PDF)"])
ws5.cell(row=1, column=1).font = header_font_white
ws5.cell(row=1, column=2).font = header_font_white
ws5.cell(row=1, column=1).fill = header_fill
ws5.cell(row=1, column=2).fill = header_fill

for folder in sorted(folder_contents.keys()):
    if folder in ('converted', 'converted_new', 'outputs', 'summary', 'temp_docx', '.'):
        continue
    for fname in sorted(folder_contents[folder]):
        ws5.append([folder, fname])

# Save
output_path = r"D:\AI_Project\Книги\summary\library_inventory.xlsx"
wb.save(output_path)
print(f"Excel saved to: {output_path}")
print("Done!")
